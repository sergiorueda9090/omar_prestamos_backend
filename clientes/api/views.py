from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from clientes.models import Cliente, Cuota
from .serializers import ClienteSerializer, CuotaSerializer
from rest_framework.pagination import PageNumberPagination
import json

from datetime import datetime
from django.db.models import Q

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse

# --- CLIENTES ---
@api_view(['GET'])
def listar_clientes(request):
    clientes = Cliente.objects.all().order_by('-created_at')

    # Inicializar el paginador
    paginator = PageNumberPagination()
    paginator.page_size = 20  # Puedes cambiar este número si quieres más o menos por página

    # Paginar la queryset
    resultado_paginado = paginator.paginate_queryset(clientes, request)

    # Serializar los datos paginados
    serializer = ClienteSerializer(resultado_paginado, many=True)

    # Retornar respuesta paginada
    return paginator.get_paginated_response(serializer.data)

@api_view(['POST'])
def crear_cliente(request):
    data = request.data.copy()

    cuotas_raw = data.pop('cuotas', [])

    # Asegurarse de que cuotas_raw sea string antes de hacer json.loads
    if isinstance(cuotas_raw, list) and len(cuotas_raw) == 1:
        cuotas_raw = cuotas_raw[0]  # Extraer el string del array

    # Parsear el JSON
    cuotas_data = json.loads(cuotas_raw)

    # Crear cliente
    cliente_serializer = ClienteSerializer(data=data)
    if cliente_serializer.is_valid():
        cliente = cliente_serializer.save()

        for cuota in cuotas_data:
            cuota['cliente'] = cliente.id
            cuota_serializer = CuotaSerializer(data=cuota)
            if cuota_serializer.is_valid():
                cuota_serializer.save()
            else:
                cliente.delete()
                return Response({
                    'error': 'Error al crear una de las cuotas',
                    'detalle': cuota_serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)

        return Response(cliente_serializer.data, status=status.HTTP_201_CREATED)

    return Response(cliente_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def obtener_cliente(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ClienteSerializer(cliente)
    return Response(serializer.data)

@api_view(['PUT'])
def actualizar_cliente(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ClienteSerializer(cliente, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def eliminar_cliente(request, cliente_id):
    try:
        cliente         = Cliente.objects.get(pk=cliente_id)
        total_cuotas    = cliente.cuotas.count()
        cliente.delete()

        return Response({
            'mensaje': 'Cliente y cuotas eliminados correctamente',
            'cliente_id': cliente_id,
            'cuotas_eliminadas': total_cuotas
        })

    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
def exportar_clientes_excel(request):
    search = request.GET.get('search', '')
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    clientes = Cliente.objects.all().order_by('-created_at')

    if search:
        clientes = clientes.filter(
            Q(nombre__icontains=search) |
            Q(numero_tarjeta__icontains=search)
        )

    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            clientes = clientes.filter(fecha_prestamo__date__gte=start)
        except ValueError:
            return Response({'error': 'startDate inválida. Usa formato YYYY-MM-DD.'}, status=400)

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            clientes = clientes.filter(fecha_prestamo__date__lte=end)
        except ValueError:
            return Response({'error': 'endDate inválida. Usa formato YYYY-MM-DD.'}, status=400)

    clientes = clientes.prefetch_related('cuotas')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes con Cuotas"

    headers = [
        "ID Cliente",
        "Nombre completo",
        "Número de tarjeta",
        "Fecha de préstamo",
        "Día de cobro",
        "Tipo de préstamo",
        "Duración (meses)",
        "Monto del préstamo",
        "Porcentaje de interés (%)",
        "Interés mensual",
        "Valor de cada cuota",
        "Número total de cuotas",
        "Total a pagar en intereses",
        "Saldo total a pagar",
        "Fecha de pago de cuota",
        "Valor de cuota",
        "Estado de pago"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cliente in clientes:
        cuotas = cliente.cuotas.all().order_by('fecha_pago')

        if cuotas.exists():
            for cuota in cuotas:
                ws.append([
                    cliente.id,
                    cliente.nombre,
                    cliente.numero_tarjeta,
                    cliente.fecha_prestamo.strftime("%Y-%m-%d"),
                    cliente.dia_cobro.strftime("%Y-%m-%d"),
                    cliente.tipo_prestamo,
                    cliente.duracion_prestamo,
                    cliente.monto_prestamo,
                    cliente.porcentaje_interes,
                    cliente.interes_mensual,
                    cliente.valor_cuota,
                    cliente.numero_cuotas,
                    cliente.total_interes_pagar,
                    cliente.saldo_total_pagar,
                    cuota.fecha_pago.strftime("%Y-%m-%d"),
                    cuota.valor,
                    cuota.estado_pago
                ])
        else:
            ws.append([
                cliente.id,
                cliente.nombre,
                cliente.numero_tarjeta,
                cliente.fecha_prestamo.strftime("%Y-%m-%d"),
                cliente.dia_cobro.strftime("%Y-%m-%d"),
                cliente.tipo_prestamo,
                cliente.duracion_prestamo,
                cliente.monto_prestamo,
                cliente.porcentaje_interes,
                cliente.interes_mensual,
                cliente.valor_cuota,
                cliente.numero_cuotas,
                cliente.total_interes_pagar,
                cliente.saldo_total_pagar,
                "", "", ""
            ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=clientes_con_cuotas.xlsx'
    wb.save(response)
    return response

# --- CUOTAS ---
@api_view(['GET'])
def listar_cuotas(request):
    cuotas = Cuota.objects.all().order_by('fecha_pago')
    serializer = CuotaSerializer(cuotas, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def crear_cuota(request):
    serializer = CuotaSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
def actualizar_estado_cuota(request, cuota_id):
    try:
        cuota = Cuota.objects.get(pk=cuota_id)
    except Cuota.DoesNotExist:
        return Response({'error': 'Cuota no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    serializer = CuotaSerializer(cuota, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
