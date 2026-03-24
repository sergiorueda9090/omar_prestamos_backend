from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from clientes.models import Cliente, Cuota, Pago, PagoInteres, HistorialEvento, Ampliacion
from .serializers import (
    ClienteSerializer, CuotaSerializer,
    PagoSerializer, PagoInteresSerializer, HistorialEventoSerializer,
    AmpliacionSerializer, ClienteDetalleCompletoSerializer,
)
from rest_framework.pagination import PageNumberPagination
import json

from datetime import datetime
from django.db.models import Q
from django.db import transaction

from clientes.utils import (
    calcular_interes_simple,
    calcular_numero_cuotas,
    calcular_fechas_cobro,
    aplicar_saldo_favor,
    distribuir_pago_en_cuotas,
    format_money,
    parse_money,
)

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse

# --- CLIENTES ---
@api_view(['GET'])
def listar_clientes(request):
    clientes = Cliente.objects.all().order_by('-created_at')

    # Inicializar el paginador
    paginator = PageNumberPagination()
    paginator.page_size = 20  # Puedes cambiar este número si quieres más o menos por página

    # Paginar la queryset
    resultado_paginado = paginator.paginate_queryset(clientes, request)

    # Serializar los datos paginados
    serializer = ClienteSerializer(resultado_paginado, many=True)

    # Retornar respuesta paginada
    return paginator.get_paginated_response(serializer.data)

@api_view(['POST'])
def crear_cliente(request):
    data = request.data.copy()

    cuotas_raw = data.pop('cuotas', [])

    # Asegurarse de que cuotas_raw sea string antes de hacer json.loads
    if isinstance(cuotas_raw, list) and len(cuotas_raw) == 1:
        cuotas_raw = cuotas_raw[0]  # Extraer el string del array

    # Parsear el JSON
    cuotas_data = json.loads(cuotas_raw)

    # Crear cliente
    cliente_serializer = ClienteSerializer(data=data)
    if cliente_serializer.is_valid():
        cliente = cliente_serializer.save()

        for cuota in cuotas_data:
            cuota['cliente'] = cliente.id
            cuota_serializer = CuotaSerializer(data=cuota)
            if cuota_serializer.is_valid():
                cuota_serializer.save()
            else:
                cliente.delete()
                return Response({
                    'error': 'Error al crear una de las cuotas',
                    'detalle': cuota_serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)

        return Response(cliente_serializer.data, status=status.HTTP_201_CREATED)

    return Response(cliente_serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def obtener_cliente(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ClienteSerializer(cliente)
    return Response(serializer.data)

@api_view(['PUT'])
def actualizar_cliente(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ClienteSerializer(cliente, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def eliminar_cliente(request, cliente_id):
    try:
        cliente         = Cliente.objects.get(pk=cliente_id)
        total_cuotas    = cliente.cuotas.count()
        cliente.delete()

        return Response({
            'mensaje': 'Cliente y cuotas eliminados correctamente',
            'cliente_id': cliente_id,
            'cuotas_eliminadas': total_cuotas
        })

    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['GET'])
def exportar_clientes_excel(request):
    search = request.GET.get('search', '')
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')

    clientes = Cliente.objects.all().order_by('-created_at')

    if search:
        clientes = clientes.filter(
            Q(nombre__icontains=search) |
            Q(numero_tarjeta__icontains=search)
        )

    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            clientes = clientes.filter(fecha_prestamo__date__gte=start)
        except ValueError:
            return Response({'error': 'startDate inválida. Usa formato YYYY-MM-DD.'}, status=400)

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            clientes = clientes.filter(fecha_prestamo__date__lte=end)
        except ValueError:
            return Response({'error': 'endDate inválida. Usa formato YYYY-MM-DD.'}, status=400)

    clientes = clientes.prefetch_related('cuotas')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes con Cuotas"

    headers = [
        "ID Cliente",
        "Nombre completo",
        "Número de tarjeta",
        "Fecha de préstamo",
        "Día de cobro",
        "Tipo de préstamo",
        "Duración (meses)",
        "Monto del préstamo",
        "Porcentaje de interés (%)",
        "Interés mensual",
        "Valor de cada cuota",
        "Número total de cuotas",
        "Total a pagar en intereses",
        "Saldo total a pagar",
        "Fecha de pago de cuota",
        "Valor de cuota",
        "Estado de pago"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cliente in clientes:
        cuotas = cliente.cuotas.all().order_by('fecha_pago')

        if cuotas.exists():
            for cuota in cuotas:
                ws.append([
                    cliente.id,
                    cliente.nombre,
                    cliente.numero_tarjeta,
                    cliente.fecha_prestamo.strftime("%Y-%m-%d"),
                    cliente.dia_cobro.strftime("%Y-%m-%d"),
                    cliente.tipo_prestamo,
                    cliente.duracion_prestamo,
                    cliente.monto_prestamo,
                    cliente.porcentaje_interes,
                    cliente.interes_mensual,
                    cliente.valor_cuota,
                    cliente.numero_cuotas,
                    cliente.total_interes_pagar,
                    cliente.saldo_total_pagar,
                    cuota.fecha_pago.strftime("%Y-%m-%d"),
                    cuota.valor,
                    cuota.estado_pago
                ])
        else:
            ws.append([
                cliente.id,
                cliente.nombre,
                cliente.numero_tarjeta,
                cliente.fecha_prestamo.strftime("%Y-%m-%d"),
                cliente.dia_cobro.strftime("%Y-%m-%d"),
                cliente.tipo_prestamo,
                cliente.duracion_prestamo,
                cliente.monto_prestamo,
                cliente.porcentaje_interes,
                cliente.interes_mensual,
                cliente.valor_cuota,
                cliente.numero_cuotas,
                cliente.total_interes_pagar,
                cliente.saldo_total_pagar,
                "", "", ""
            ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=clientes_con_cuotas.xlsx'
    wb.save(response)
    return response

# --- CUOTAS ---
@api_view(['GET'])
def listar_cuotas(request):
    cuotas = Cuota.objects.all().order_by('fecha_pago')
    serializer = CuotaSerializer(cuotas, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def crear_cuota(request):
    serializer = CuotaSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['PUT'])
def actualizar_estado_cuota(request, cuota_id):
    try:
        cuota = Cuota.objects.get(pk=cuota_id)
    except Cuota.DoesNotExist:
        return Response({'error': 'Cuota no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    serializer = CuotaSerializer(cuota, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# =============================================================================
# =============================================================================
#                          ENDPOINTS V2
#          (Para ClientesTestMejorado — todo se persiste en DB)
# =============================================================================
# =============================================================================


# =============================================================================
# 1. LISTAR CLIENTES v2 (paginado)
# GET /clientes/api/v2/
# =============================================================================

@api_view(['GET'])
def listar_clientes_v2(request):
    clientes = Cliente.objects.all().order_by('-created_at')

    # --- Filtros opcionales ---
    search = request.GET.get('search', '')
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')
    estado = request.GET.get('estado')

    if estado:
        clientes = clientes.filter(estado=estado)

    if search:
        clientes = clientes.filter(
            Q(nombre__icontains=search) |
            Q(numero_tarjeta__icontains=search)
        )

    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            clientes = clientes.filter(fecha_prestamo__gte=start)
        except ValueError:
            pass

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            clientes = clientes.filter(fecha_prestamo__lte=end)
        except ValueError:
            pass

    paginator = PageNumberPagination()
    paginator.page_size = 20

    resultado_paginado = paginator.paginate_queryset(clientes, request)
    serializer = ClienteSerializer(resultado_paginado, many=True)

    return paginator.get_paginated_response(serializer.data)


# =============================================================================
# 2. CREAR PRESTAMO v2
# POST /clientes/api/v2/crear/
#
# Body (JSON):
# {
#   "numero_tarjeta": "TC-123456",
#   "nombre": "Juan Perez",
#   "monto_prestamo": "1000000",
#   "porcentaje_interes": "10",
#   "duracion_prestamo": 12,
#   "tipo_prestamo": "Mensual",
#   "fecha_prestamo": "2026-03-23",
#   "dia_cobro": "2026-04-23",
#   "prestamo_sin_cronograma": false
# }
#
# El backend calcula: interes_mensual, numero_cuotas, valor_cuota,
# total_interes_pagar, saldo_total_pagar.
# Si NO es sin cronograma, genera las cuotas automaticamente.
# =============================================================================

@api_view(['POST'])
@transaction.atomic
def crear_cliente_v2(request):
    data = request.data.copy()

    # --- Extraer datos basicos ---
    monto = parse_money(data.get('monto_prestamo', '0'))
    tasa = float(data.get('porcentaje_interes', '0'))
    duracion = int(data.get('duracion_prestamo', '0'))
    tipo_prestamo = data.get('tipo_prestamo', 'Mensual')
    sin_cronograma = data.get('prestamo_sin_cronograma', False)
    fecha_prestamo = data.get('fecha_prestamo', '')
    dia_cobro = data.get('dia_cobro', '')

    # --- Validaciones ---
    if not monto or not tasa or not duracion:
        return Response(
            {'error': 'Monto, tasa e interés y duración son obligatorios'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # --- Calcular valores del prestamo ---
    num_cuotas = calcular_numero_cuotas(duracion, tipo_prestamo)
    calculo = calcular_interes_simple(monto, tasa, num_cuotas, duracion)

    # --- Preparar datos para el serializer ---
    data['interes_mensual'] = str(int(calculo['interes_mensual']))
    data['numero_cuotas'] = num_cuotas
    data['valor_cuota'] = str(int(calculo['valor_cuota']))
    data['total_interes_pagar'] = str(int(calculo['total_interes']))
    data['saldo_total_pagar'] = str(int(calculo['saldo_total']))
    data['prestamo_sin_cronograma'] = sin_cronograma
    data['estado'] = 'vigente'
    data['monto_prestamo'] = str(int(monto))

    # --- Crear el cliente ---
    serializer = ClienteSerializer(data=data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    cliente = serializer.save()

    # --- Generar cuotas (solo si tiene cronograma) ---
    if not sin_cronograma:
        fechas = calcular_fechas_cobro(fecha_prestamo, num_cuotas, tipo_prestamo, dia_cobro)
        valor_cuota_str = str(int(calculo['valor_cuota']))

        for i, fecha in enumerate(fechas):
            Cuota.objects.create(
                cliente=cliente,
                numero=i + 1,
                fecha_pago=fecha,
                valor=valor_cuota_str,
                abonado='0',
                saldo=valor_cuota_str,
                estado_pago='pendiente',
            )

    # --- Crear evento en historial ---
    if sin_cronograma:
        desc = (
            f"Monto: ${format_money(monto)}, Tasa: {tasa}%, "
            f"Plazo: {duracion} meses ({num_cuotas} cuotas {tipo_prestamo})"
        )
        titulo = 'Préstamo Creado (Sin Cronograma)'
    else:
        desc = (
            f"Monto: ${format_money(monto)}, Tasa: {tasa}%, "
            f"{num_cuotas} cuotas de ${format_money(calculo['valor_cuota'])}"
        )
        titulo = 'Préstamo Creado'

    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='creacion',
        titulo=titulo,
        descripcion=desc,
        monto=str(int(monto)),
    )

    # --- Respuesta completa ---
    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data, status=status.HTTP_201_CREATED)


# =============================================================================
# 3. OBTENER DETALLE COMPLETO DE UN PRESTAMO
# GET /clientes/api/v2/<id>/
#
# Retorna el cliente con todas sus relaciones anidadas:
# cuotas, pagos, pagos_intereses, historial, ampliaciones
# =============================================================================

@api_view(['GET'])
def obtener_cliente_detalle(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(serializer.data)


# =============================================================================
# 4. REGISTRAR PAGO DE CUOTA (con cronograma)
# POST /clientes/api/v2/<id>/pagar/
#
# Body: { "monto": "150000", "fecha_pago": "2026-03-23" }
#
# Distribuye el monto entre cuotas pendientes en orden.
# Actualiza abonado, saldo, estado_pago de cada cuota afectada.
# =============================================================================

@api_view(['POST'])
@transaction.atomic
def registrar_pago(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    monto = parse_money(request.data.get('monto', '0'))
    fecha_pago = request.data.get('fecha_pago', datetime.now().strftime('%Y-%m-%d'))

    if monto <= 0:
        return Response({'error': 'El monto debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)

    # --- Distribuir pago entre cuotas pendientes ---
    cuotas = cliente.cuotas.all()
    distribucion = distribuir_pago_en_cuotas(cuotas, monto)

    if not distribucion:
        return Response({'error': 'No hay cuotas pendientes'}, status=status.HTTP_400_BAD_REQUEST)

    # --- Crear registro de pago ---
    Pago.objects.create(
        cliente=cliente,
        tipo_pago='cuota',
        monto=str(int(monto)),
        fecha=fecha_pago,
        descripcion=f"Pago distribuido en cuota(s): {', '.join(['#' + str(d['numero']) for d in distribucion])}",
    )

    # --- Verificar si todas las cuotas estan pagadas ---
    cuotas_pendientes = cliente.cuotas.exclude(estado_pago='pagado').count()
    if cuotas_pendientes == 0:
        cliente.estado = 'pagado'
        cliente.save()

    # --- Crear evento en historial ---
    cuotas_afectadas = ', '.join([f"#{d['numero']}" for d in distribucion])
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='pago',
        titulo='Pago Registrado',
        descripcion=f"Pago de ${format_money(monto)} distribuido en cuota(s): {cuotas_afectadas}",
        monto=str(int(monto)),
    )

    # --- Respuesta completa ---
    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 5. REGISTRAR PAGO SIN CRONOGRAMA
# POST /clientes/api/v2/<id>/pagar-sin-cronograma/
#
# Body: { "monto": "150000", "fecha_pago": "2026-03-23" }
#
# Incrementa total_pagado_cuotas_sin_cronograma.
# =============================================================================

@api_view(['POST'])
@transaction.atomic
def registrar_pago_sin_cronograma(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    monto = parse_money(request.data.get('monto', '0'))
    fecha_pago = request.data.get('fecha_pago', datetime.now().strftime('%Y-%m-%d'))

    if monto <= 0:
        return Response({'error': 'El monto debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)

    # --- Incrementar total pagado ---
    total_actual = parse_money(cliente.total_pagado_cuotas_sin_cronograma)
    nuevo_total = total_actual + monto
    cliente.total_pagado_cuotas_sin_cronograma = str(int(nuevo_total))

    # --- Verificar si ya pago todo ---
    saldo_total = parse_money(cliente.saldo_total_pagar)
    if nuevo_total >= saldo_total:
        cliente.estado = 'pagado'

    cliente.save()

    # --- Crear registro de pago ---
    valor_cuota = parse_money(cliente.valor_cuota)
    cuotas_equivalentes = round(monto / valor_cuota, 2) if valor_cuota > 0 else 0

    Pago.objects.create(
        cliente=cliente,
        tipo_pago='cuota_sin_cronograma',
        monto=str(int(monto)),
        fecha=fecha_pago,
        descripcion=f"Pago de ${format_money(monto)} ({cuotas_equivalentes} cuotas equivalentes)",
    )

    # --- Crear evento en historial ---
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='pago',
        titulo='Pago de Cuota (Sin Cronograma)',
        descripcion=f"Pago de ${format_money(monto)} ({cuotas_equivalentes} cuotas equivalentes). Fecha: {fecha_pago}",
        monto=str(int(monto)),
    )

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 6. REGISTRAR PAGO DE INTERES
# POST /clientes/api/v2/<id>/pagar-interes/
#
# Body: { "monto": "50000", "fecha": "2026-03-23", "descripcion": "Interés mora" }
#
# Crea PagoInteres y suma al interes_acumulado del cliente.
# =============================================================================

@api_view(['POST'])
@transaction.atomic
def registrar_pago_interes(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    monto = parse_money(request.data.get('monto', '0'))
    fecha = request.data.get('fecha', datetime.now().strftime('%Y-%m-%d'))
    descripcion = request.data.get('descripcion', '')

    if monto <= 0:
        return Response({'error': 'El monto debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)

    # --- Crear PagoInteres ---
    pago_interes = PagoInteres.objects.create(
        cliente=cliente,
        fecha=fecha,
        monto=str(int(monto)),
        tipo='interes',
        descripcion=descripcion or 'Pago de Interés',
    )

    # --- Incrementar interes acumulado ---
    interes_actual = parse_money(cliente.interes_acumulado)
    cliente.interes_acumulado = str(int(interes_actual + monto))
    cliente.save()

    # --- Crear registro de pago general ---
    Pago.objects.create(
        cliente=cliente,
        tipo_pago='interes',
        monto=str(int(monto)),
        fecha=fecha,
        descripcion=descripcion or 'Pago de Interés',
    )

    # --- Evento historial ---
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='pago',
        titulo='Pago de Interés Registrado',
        descripcion=f"Pago de interés: ${format_money(monto)}. Se suma a las 3 utilidades.{' ' + descripcion if descripcion else ''}",
        monto=str(int(monto)),
    )

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 7. ELIMINAR PAGO DE INTERES
# DELETE /clientes/api/v2/pagos-interes/<id>/
#
# Resta el monto del interes_acumulado y elimina el registro.
# =============================================================================

@api_view(['DELETE'])
@transaction.atomic
def eliminar_pago_interes(request, pago_id):
    try:
        pago = PagoInteres.objects.get(pk=pago_id)
    except PagoInteres.DoesNotExist:
        return Response({'error': 'Pago de interés no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    cliente = pago.cliente
    monto = parse_money(pago.monto)

    # --- Restar del interes acumulado ---
    interes_actual = parse_money(cliente.interes_acumulado)
    cliente.interes_acumulado = str(int(max(0, interes_actual - monto)))
    cliente.save()

    # --- Evento historial ---
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='eliminacion_pago',
        titulo='Pago de Interés Eliminado',
        descripcion=f"Se eliminó pago de interés por ${format_money(monto)}. Fecha: {pago.fecha}",
        monto=str(int(monto)),
    )

    # --- Eliminar ---
    pago.delete()

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 8. PAGAR SALDO TOTAL
# POST /clientes/api/v2/<id>/pagar-saldo-total/
#
# Body: {
#   "porcentaje_interes": "5",
#   "tiempo": "3",
#   "fecha_pago": "2026-03-23"
# }
#
# Calcula totalBruto = capital + (capital * tasa% * meses)
# Marca todas las cuotas como pagadas.
# =============================================================================

@api_view(['POST'])
@transaction.atomic
def registrar_pago_saldo_total(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    porcentaje = float(request.data.get('porcentaje_interes', '0'))
    tiempo = int(request.data.get('tiempo', '0'))
    fecha_pago = request.data.get('fecha_pago', datetime.now().strftime('%Y-%m-%d'))

    if tiempo <= 0:
        return Response({'error': 'Tiempo debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)

    dinero_prestado = parse_money(cliente.monto_prestamo)

    # --- Calcular total bruto ---
    total_bruto = dinero_prestado + (dinero_prestado * (porcentaje / 100) * tiempo)

    # --- Si es prestamo sin cronograma ---
    if cliente.prestamo_sin_cronograma:
        abono_total = parse_money(cliente.total_pagado_cuotas_sin_cronograma)
        total_a_pagar = max(0, total_bruto - abono_total)

        # Actualizar cliente
        cliente.total_pagado_cuotas_sin_cronograma = str(int(total_bruto))
        cliente.saldo_total_pagar = str(int(total_bruto))
        cliente.estado = 'pagado'
        cliente.save()
    else:
        # --- Calcular abono total de cuotas ---
        abono_total = sum(parse_money(c.abonado) for c in cliente.cuotas.all())
        total_a_pagar = max(0, total_bruto - abono_total)

        # --- Redistribuir total_bruto entre todas las cuotas ---
        cuotas = cliente.cuotas.all()
        num_cuotas = cuotas.count()
        if num_cuotas > 0:
            nuevo_valor_cuota = total_bruto / num_cuotas
            for cuota in cuotas:
                cuota.valor = str(int(nuevo_valor_cuota))
                cuota.abonado = str(int(nuevo_valor_cuota))
                cuota.saldo = '0'
                cuota.estado_pago = 'pagado'
                cuota.save()

        # Actualizar datos del prestamo
        cliente.saldo_total_pagar = str(int(total_bruto))
        cliente.total_interes_pagar = str(int(total_bruto - dinero_prestado))
        if num_cuotas > 0:
            cliente.valor_cuota = str(int(total_bruto / num_cuotas))
        cliente.estado = 'pagado'
        cliente.save()

    # --- Crear registro de pago ---
    Pago.objects.create(
        cliente=cliente,
        tipo_pago='saldo_total',
        monto=str(int(total_a_pagar)),
        fecha=fecha_pago,
        descripcion=f"Pago saldo total. Bruto: ${format_money(total_bruto)}, Abono previo: ${format_money(abono_total)}",
    )

    # --- Evento historial ---
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='pago',
        titulo='Pago de Saldo Total Registrado',
        descripcion=(
            f"Saldo total pagado: ${format_money(total_bruto)}. "
            f"Interés: {porcentaje}%, Tiempo: {tiempo} meses. Fecha: {fecha_pago}"
        ),
        monto=str(int(total_a_pagar)),
    )

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 9. CAMBIAR FECHA DE CUOTA
# PUT /clientes/api/v2/cuotas/<id>/cambiar-fecha/
#
# Body: { "fecha_pago": "2026-05-15" }
# =============================================================================

@api_view(['PUT'])
@transaction.atomic
def cambiar_fecha_cuota(request, cuota_id):
    try:
        cuota = Cuota.objects.get(pk=cuota_id)
    except Cuota.DoesNotExist:
        return Response({'error': 'Cuota no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    nueva_fecha = request.data.get('fecha_pago')
    if not nueva_fecha:
        return Response({'error': 'fecha_pago es obligatorio'}, status=status.HTTP_400_BAD_REQUEST)

    fecha_anterior = str(cuota.fecha_pago)
    cuota.fecha_pago = nueva_fecha
    cuota.save()

    # --- Evento historial ---
    HistorialEvento.objects.create(
        cliente=cuota.cliente,
        tipo='cambio_fecha',
        titulo=f'Fecha Modificada - Cuota #{cuota.numero}',
        descripcion=f"Fecha anterior: {fecha_anterior}, Nueva fecha: {nueva_fecha}",
    )

    response_serializer = ClienteDetalleCompletoSerializer(cuota.cliente)
    return Response(response_serializer.data)


# =============================================================================
# 10. ELIMINAR/REVERTIR PAGO DE UNA CUOTA
# POST /clientes/api/v2/cuotas/<id>/eliminar-pago/
#
# Resetea abonado=0, saldo=valor, estado_pago='pendiente'
# =============================================================================

@api_view(['POST'])
@transaction.atomic
def eliminar_pago_cuota(request, cuota_id):
    try:
        cuota = Cuota.objects.get(pk=cuota_id)
    except Cuota.DoesNotExist:
        return Response({'error': 'Cuota no encontrada'}, status=status.HTTP_404_NOT_FOUND)

    monto_eliminado = parse_money(cuota.abonado)
    if monto_eliminado == 0:
        return Response({'error': 'Esta cuota no tiene pagos para eliminar'}, status=status.HTTP_400_BAD_REQUEST)

    # --- Resetear cuota ---
    cuota.abonado = '0'
    cuota.saldo = cuota.valor
    cuota.estado_pago = 'pendiente'
    cuota.save()

    # --- Si el prestamo estaba pagado, volver a vigente ---
    cliente = cuota.cliente
    if cliente.estado == 'pagado':
        cliente.estado = 'vigente'
        cliente.save()

    # --- Evento historial ---
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='eliminacion_pago',
        titulo=f'Pago Eliminado - Cuota #{cuota.numero}',
        descripcion=f"Se eliminaron pagos por ${format_money(monto_eliminado)}. La cuota volvió a estado pendiente.",
        monto=str(int(monto_eliminado)),
    )

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 11. AMPLIAR PRESTAMO
# POST /clientes/api/v2/<id>/ampliar/
#
# Body: {
#   "monto_adicional": "500000",
#   "nueva_tasa": "5",
#   "nuevas_cuotas": 12
# }
#
# Proceso:
# 1. Calcular interes de liquidacion y saldo a favor
# 2. Eliminar cuotas viejas
# 3. Generar nuevo cronograma con capital ampliado
# 4. Aplicar saldo a favor a primeras cuotas
# =============================================================================

@api_view(['POST'])
@transaction.atomic
def ampliar_prestamo(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    monto_adicional = parse_money(request.data.get('monto_adicional', '0'))
    nueva_tasa = float(request.data.get('nueva_tasa', '0'))
    nuevas_cuotas_num = int(request.data.get('nuevas_cuotas', '0'))

    if monto_adicional <= 0 or nueva_tasa <= 0 or nuevas_cuotas_num <= 0:
        return Response(
            {'error': 'Monto adicional, tasa y número de cuotas deben ser mayores a 0'},
            status=status.HTTP_400_BAD_REQUEST
        )

    capital_anterior = parse_money(cliente.monto_prestamo)

    # --- Paso 1: Calcular liquidacion ---
    if cliente.prestamo_sin_cronograma:
        # Para sin cronograma: usar cuotas proporcionales
        total_pagado = parse_money(cliente.total_pagado_cuotas_sin_cronograma)
        valor_cuota_original = parse_money(cliente.valor_cuota)
        cuotas_pagadas = total_pagado / valor_cuota_original if valor_cuota_original > 0 else 0
        cuotas_restantes = max(0, int(cliente.numero_cuotas - cuotas_pagadas))
    else:
        # Para con cronograma: contar cuotas con pagos
        total_pagado = sum(parse_money(c.abonado) for c in cliente.cuotas.filter(estado_pago__in=['pagado', 'parcial']))
        cuotas_restantes = cliente.cuotas.exclude(estado_pago='pagado').count()

    # Interes de liquidacion = capital * tasa * cuotas_restantes
    tasa_original = float(cliente.porcentaje_interes)
    liquidacion = calcular_interes_simple(capital_anterior, tasa_original, cuotas_restantes)
    interes_liquidacion = liquidacion['total_interes']

    # Saldo a favor = lo que el cliente ya pago - interes de liquidacion
    saldo_favor = total_pagado - interes_liquidacion

    # --- Paso 2: Eliminar cuotas viejas ---
    cliente.cuotas.all().delete()

    # --- Paso 3: Calcular nuevo prestamo ---
    nuevo_capital = capital_anterior + monto_adicional
    calculo = calcular_interes_simple(nuevo_capital, nueva_tasa, nuevas_cuotas_num, nuevas_cuotas_num)

    # --- Paso 4: Generar nuevo cronograma ---
    fechas = calcular_fechas_cobro(
        datetime.now().strftime('%Y-%m-%d'),
        nuevas_cuotas_num,
        cliente.tipo_prestamo,
        str(cliente.dia_cobro),
    )

    cuotas_data = []
    for i, fecha in enumerate(fechas):
        cuotas_data.append({
            'numero': i + 1,
            'fecha_pago': fecha,
            'valor': str(int(calculo['valor_cuota'])),
        })

    # Aplicar saldo a favor a las primeras cuotas
    if saldo_favor > 0:
        cuotas_data = aplicar_saldo_favor(cuotas_data, saldo_favor)
    else:
        for cuota_d in cuotas_data:
            cuota_d['abonado'] = '0'
            cuota_d['saldo'] = cuota_d['valor']
            cuota_d['estado_pago'] = 'pendiente'

    # Crear cuotas en DB
    for cuota_d in cuotas_data:
        Cuota.objects.create(
            cliente=cliente,
            numero=cuota_d['numero'],
            fecha_pago=cuota_d['fecha_pago'],
            valor=cuota_d['valor'],
            abonado=cuota_d.get('abonado', '0'),
            saldo=cuota_d.get('saldo', cuota_d['valor']),
            estado_pago=cuota_d.get('estado_pago', 'pendiente'),
        )

    # --- Paso 5: Actualizar datos del cliente ---
    cliente.monto_prestamo = str(int(nuevo_capital))
    cliente.porcentaje_interes = str(nueva_tasa)
    cliente.numero_cuotas = nuevas_cuotas_num
    cliente.duracion_prestamo = nuevas_cuotas_num
    cliente.valor_cuota = str(int(calculo['valor_cuota']))
    cliente.saldo_total_pagar = str(int(calculo['saldo_total']))
    cliente.total_interes_pagar = str(int(calculo['total_interes']))
    cliente.interes_mensual = str(int(nuevo_capital * nueva_tasa / 100))
    cliente.estado = 'vigente'
    cliente.prestamo_sin_cronograma = False  # Ampliacion siempre genera cronograma

    # Sumar interes de liquidacion al interes acumulado
    interes_acum = parse_money(cliente.interes_acumulado)
    cliente.interes_acumulado = str(int(interes_acum + interes_liquidacion))

    cliente.save()

    # --- Crear PagoInteres de liquidacion ---
    PagoInteres.objects.create(
        cliente=cliente,
        fecha=datetime.now().strftime('%Y-%m-%d'),
        monto=str(int(interes_liquidacion)),
        tipo='liquidacion',
        descripcion='Interés de Liquidación',
    )

    # --- Crear Ampliacion ---
    Ampliacion.objects.create(
        cliente=cliente,
        monto_adicional=str(int(monto_adicional)),
        nueva_tasa=str(nueva_tasa),
        nuevas_cuotas=nuevas_cuotas_num,
        capital_anterior=str(int(capital_anterior)),
        capital_nuevo=str(int(nuevo_capital)),
        interes_liquidacion=str(int(interes_liquidacion)),
        saldo_favor=str(int(max(0, saldo_favor))),
    )

    # --- Eventos historial ---
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='liquidacion',
        titulo='Liquidación Realizada',
        descripcion=(
            f"Interés de liquidación: ${format_money(interes_liquidacion)}, "
            f"Saldo a favor: ${format_money(max(0, saldo_favor))}."
        ),
        monto=str(int(interes_liquidacion)),
    )
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='ampliacion',
        titulo='Ampliación de Préstamo',
        descripcion=(
            f"Monto adicional: ${format_money(monto_adicional)}, "
            f"Nuevo total: ${format_money(nuevo_capital)}, "
            f"{nuevas_cuotas_num} cuotas"
        ),
        monto=str(int(monto_adicional)),
    )

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 12. CAMBIAR PLAZO (solo sin cronograma)
# PUT /clientes/api/v2/<id>/cambiar-plazo/
#
# Body: { "duracion_prestamo": 6 }
#
# Recalcula: numero_cuotas, valor_cuota, total_interes, saldo_total, interes_mensual
# =============================================================================

@api_view(['PUT'])
@transaction.atomic
def cambiar_plazo(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    if not cliente.prestamo_sin_cronograma:
        return Response(
            {'error': 'Solo se puede cambiar el plazo en préstamos sin cronograma'},
            status=status.HTTP_400_BAD_REQUEST
        )

    nueva_duracion = int(request.data.get('duracion_prestamo', '0'))
    if nueva_duracion <= 0:
        return Response({'error': 'La duración debe ser mayor a 0'}, status=status.HTTP_400_BAD_REQUEST)

    plazo_anterior = cliente.duracion_prestamo

    # --- Recalcular ---
    capital = parse_money(cliente.monto_prestamo)
    tasa = float(cliente.porcentaje_interes)
    # Para sin cronograma, plazo = meses = cuotas
    calculo = calcular_interes_simple(capital, tasa, nueva_duracion, nueva_duracion)

    # --- Actualizar cliente ---
    cliente.duracion_prestamo = nueva_duracion
    cliente.numero_cuotas = nueva_duracion
    cliente.valor_cuota = str(int(calculo['valor_cuota']))
    cliente.total_interes_pagar = str(int(calculo['total_interes']))
    cliente.saldo_total_pagar = str(int(calculo['saldo_total']))
    cliente.interes_mensual = str(int(calculo['interes_mensual']))
    cliente.save()

    # --- Evento historial ---
    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='cambio_plazo',
        titulo='Plazo Modificado',
        descripcion=(
            f"Plazo anterior: {plazo_anterior} meses, "
            f"Nuevo plazo: {nueva_duracion} meses. "
            f"Nueva cuota: ${format_money(calculo['valor_cuota'])}, "
            f"Nuevo total: ${format_money(calculo['saldo_total'])}"
        ),
    )

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 14. MARCAR COMO PERDIDO
# POST /clientes/api/v2/<id>/marcar-perdido/
# =============================================================================

@api_view(['POST'])
def marcar_perdido(request, cliente_id):
    try:
        cliente = Cliente.objects.get(pk=cliente_id)
    except Cliente.DoesNotExist:
        return Response({'error': 'Cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    if cliente.estado == 'perdido':
        return Response({'error': 'Este cliente ya está marcado como perdido'}, status=status.HTTP_400_BAD_REQUEST)

    estado_anterior = cliente.estado
    cliente.estado = 'perdido'
    cliente.save()

    HistorialEvento.objects.create(
        cliente=cliente,
        tipo='cambio_estado',
        titulo='Marcado como Perdido',
        descripcion=f"Estado anterior: {estado_anterior}. El cliente fue marcado como perdido.",
    )

    response_serializer = ClienteDetalleCompletoSerializer(cliente)
    return Response(response_serializer.data)


# =============================================================================
# 15. EXPORTAR REPORTE COMPLETO A EXCEL (v2)
# GET /clientes/api/v2/exportar/
# Soporta filtros: ?search=&startDate=&endDate=&estado=
# =============================================================================

@api_view(['GET'])
def exportar_clientes_excel_v2(request):
    clientes = Cliente.objects.all().order_by('-created_at')

    # --- Filtros ---
    search = request.GET.get('search', '')
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')
    estado = request.GET.get('estado')

    if estado:
        clientes = clientes.filter(estado=estado)
    if search:
        clientes = clientes.filter(
            Q(nombre__icontains=search) | Q(numero_tarjeta__icontains=search)
        )
    if start_date:
        try:
            clientes = clientes.filter(fecha_prestamo__gte=datetime.strptime(start_date, "%Y-%m-%d"))
        except ValueError:
            pass
    if end_date:
        try:
            clientes = clientes.filter(fecha_prestamo__lte=datetime.strptime(end_date, "%Y-%m-%d"))
        except ValueError:
            pass

    clientes = clientes.prefetch_related('cuotas', 'pagos_intereses')

    wb = openpyxl.Workbook()

    # --- Estilos ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    estado_fills = {
        'vigente': PatternFill(start_color="D6EFFF", end_color="D6EFFF", fill_type="solid"),
        'pagado':  PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        'perdido': PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
    }
    cuota_estado_fills = {
        'pendiente': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'parcial':   PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid"),
        'pagado':    PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    }

    def apply_header_style(ws, headers):
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
                cell.border = thin_border
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    def fmt(value):
        """Formatea valores monetarios tipo string a numero legible."""
        try:
            return f"${int(float(str(value).replace(',', ''))):,.0f}"
        except (ValueError, TypeError):
            return str(value) if value else "$0"

    # =====================================================================
    # HOJA 1: RESUMEN DE CLIENTES
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Resumen Clientes"

    apply_header_style(ws1, [
        "ID", "N Tarjeta", "Nombre", "Estado",
        "Monto Prestamo", "% Interes", "Interes Mensual",
        "Duracion (meses)", "Tipo", "Fecha Prestamo", "Dia Cobro",
        "N Cuotas", "Valor Cuota", "Total Intereses", "Saldo Total",
        "Sin Cronograma", "Total Pagado (S/C)", "Interes Acumulado",
        "Utilidad Real 1", "Utilidad Real 2", "Utilidad Real 3",
    ])

    for c in clientes:
        monto_original = parse_money(c.monto_prestamo)
        total_interes = parse_money(c.total_interes_pagar)
        total_a_pagar = parse_money(c.saldo_total_pagar)
        interes_acum = parse_money(c.interes_acumulado)

        cuotas_cliente = list(c.cuotas.all())
        total_pagado_cuotas = sum(parse_money(cuota.abonado) for cuota in cuotas_cliente)

        # Utilidad 1: (totalAPagar - montoOriginal) + interesAcumulado
        utilidad1 = (total_a_pagar - monto_original) + interes_acum

        # Utilidad 2: (totalPagado * proporcionInteres) + interesAcumulado
        if total_a_pagar > 0:
            proporcion_interes = total_interes / total_a_pagar
            utilidad2 = (total_pagado_cuotas * proporcion_interes) + interes_acum
        else:
            utilidad2 = interes_acum

        # Utilidad 3: si totalPagado > montoOriginal => (totalPagado - montoOriginal) + interesAcumulado, sino 0
        if total_pagado_cuotas > monto_original:
            utilidad3 = (total_pagado_cuotas - monto_original) + interes_acum
        else:
            utilidad3 = 0

        row = ws1.max_row + 1
        ws1.append([
            c.id, c.numero_tarjeta, c.nombre, c.estado.upper(),
            fmt(c.monto_prestamo), f"{c.porcentaje_interes}%", fmt(c.interes_mensual),
            c.duracion_prestamo, c.tipo_prestamo,
            c.fecha_prestamo.strftime("%Y-%m-%d"), c.dia_cobro.strftime("%Y-%m-%d"),
            c.numero_cuotas, fmt(c.valor_cuota), fmt(c.total_interes_pagar), fmt(c.saldo_total_pagar),
            "Si" if c.prestamo_sin_cronograma else "No",
            fmt(c.total_pagado_cuotas_sin_cronograma), fmt(c.interes_acumulado),
            fmt(utilidad1), fmt(utilidad2), fmt(int(utilidad3)),
        ])
        fill = estado_fills.get(c.estado)
        if fill:
            for cell in ws1[row]:
                cell.fill = fill

    auto_width(ws1)

    # =====================================================================
    # HOJA 2: DETALLE DE CUOTAS
    # =====================================================================
    ws2 = wb.create_sheet("Detalle Cuotas")

    apply_header_style(ws2, [
        "ID Cliente", "N Tarjeta", "Nombre", "Estado Prestamo",
        "Cuota #", "Fecha Pago", "Valor Cuota", "Abonado", "Saldo Cuota", "Estado Cuota",
    ])

    for c in clientes:
        cuotas = c.cuotas.all().order_by('numero')
        for cuota in cuotas:
            row = ws2.max_row + 1
            ws2.append([
                c.id, c.numero_tarjeta, c.nombre, c.estado.upper(),
                cuota.numero, cuota.fecha_pago.strftime("%Y-%m-%d"),
                fmt(cuota.valor), fmt(cuota.abonado), fmt(cuota.saldo),
                cuota.estado_pago.upper(),
            ])
            fill = cuota_estado_fills.get(cuota.estado_pago)
            if fill:
                for cell in ws2[row]:
                    cell.fill = fill

    auto_width(ws2)

    # =====================================================================
    # HOJA 3: PAGOS DE INTERES
    # =====================================================================
    ws3 = wb.create_sheet("Pagos de Interes")

    apply_header_style(ws3, [
        "ID Cliente", "N Tarjeta", "Nombre", "Estado Prestamo",
        "Fecha Pago", "Monto", "Tipo", "Descripcion",
    ])

    for c in clientes:
        for pi in c.pagos_intereses.all().order_by('fecha'):
            ws3.append([
                c.id, c.numero_tarjeta, c.nombre, c.estado.upper(),
                pi.fecha.strftime("%Y-%m-%d"), fmt(pi.monto),
                pi.tipo.upper(), pi.descripcion,
            ])

    auto_width(ws3)

    # --- Respuesta ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=reporte_clientes_completo.xlsx'
    wb.save(response)
    return response
